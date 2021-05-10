import re
import os
from hashlib import blake2b
from more_itertools import chunked
from openpyxl import load_workbook
from datetime import datetime
from .database.utils import session_scope
from .database.models import Violation, Accident, CountryTown


uid_pattern = re.compile('^[A-Z][0-9]{9}$')


class DataMissingError(Exception):
    """Uid missing from source."""
    def __init__(self, fields):
        super().__init__(f"DataMissingError: fields={','.join(fields)}")


class DataValidationError(Exception):
    """Invalid data."""
    def __init__(self, field, msg):
        super().__init__(f"DataValidationError: field={field} | msg={msg}")


def validate_uid(uid):
    return bool(re.match(uid_pattern, str(uid)))


def violation_row_to_instance(row):
    uid, birthday, date = map(lambda idx: row[idx].value, [3, 4, 16])
    invalid_fields = []
    if birthday == 'NULL':
        birthday = None
    if not uid:
        invalid_fields.append('uid')
    if not date:
        invalid_fields.append('date')
    if invalid_fields:
        raise DataMissingError(invalid_fields)
    if not validate_uid(uid):
        raise DataValidationError('uid', f'invalid value {uid}')
    return Violation(
        uid=uid,
        birthday=datetime.strptime(birthday, '%m/%d/%Y').date() if birthday else None,
        date=date.date() if isinstance(date, datetime) else datetime.strptime(date, '%Y-%m-%d'),
    )


def accident_row_to_instance(row):
    meta = [
        ('date', 0, datetime, True), ('hour', 1, int, True), ('country', 2, str, True),
        ('town', 3, str, True),  ('dead', 6, int, True), ('injured', 7, int, True),
        ('driver_birthday', 10, str, False),
        ('party_no', 11, int, True), ('gender_no', 12, int, True), ('uid', 13, str, True),
        ('birthday', 15, str, True), ('party_address_code', 17, int, False),
        ('party_country', 18, str, False), ('party_town', 19, str, False),
        ('injury_level', 21, int, True), ('driver_level', 22, int, True),
        ('drunk_level', 23, int, True), ('hit_and_run', 24, int, True), ('job', 25, int, True),
        ('car_type', 26, str, True), ('cause_type', 27, int, False),
    ]
    fields = dict()
    for field, idx, _type, required in meta:
        value = row[idx].value
        if type(value) == str:
            value = value.strip()
        if value in ('', 'NULL'):
            value = None
        if required and value is None:
            raise DataMissingError([field])
        if value is not None and not type(value) == _type:
            raise DataValidationError(field, f"'{value}' is not {_type}")
        fields[field] = value

    if not validate_uid(fields['uid']):
        raise DataValidationError('uid', f"invalid value {fields['uid']}")

    format_driver_birthday = None
    if fields['driver_birthday']:
        try:
            format_driver_birthday = datetime.strptime(fields['driver_birthday'], '%m/%d/%Y').date()
        except ValueError:
            exc = DataValidationError('driver_birthday', f"invalid value {fields['driver_birthday']}")
            print(exc)

    format_birthday = None
    try:
        format_birthday = datetime.strptime(fields['birthday'], '%Y/%m/%d').date()
    except ValueError:
        exc = DataValidationError('birthday', f"invalid value {fields['birthday']}")
        print(exc)

    return Accident(
        date=fields['date'].date(),
        hour=fields['hour'],
        dead=fields['dead'],
        injured=fields['injured'],
        party_no=fields['party_no'],
        gender_no=fields['gender_no'],
        uid=fields['uid'],
        driver_birthday=format_driver_birthday,
        birthday=format_birthday,
        country=fields['country'],
        town=fields['town'],
        party_address_code=fields['party_address_code'],
        party_country=fields['party_country'],
        party_town=fields['party_town'],
        injury_level=fields['injury_level'],
        driver_level=fields['driver_level'],
        drunk_level=fields['drunk_level'],
        hit_and_run=fields['hit_and_run'],
        job=fields['job'],
        cause_type=fields['cause_type'],
        car_type=fields['car_type'],
    )


def region_data_to_instances(row, mapping, model_class):
    instances = []
    address_code = row[2].value
    for idx, year in mapping.items():
        rate = row[idx].value
        if not address_code:
            raise DataMissingError(fields=['address_code'])
        if rate is None:
            raise DataMissingError(fields=['rate'])
        try:
            float(rate)
        except ValueError:
            raise DataValidationError(field='rate', msg=f'Invalid value {rate}')
        instances.append(
            model_class(
                address_code=address_code,
                year=year,
                rate=rate
            )
        )
    return instances


def country_town_data_to_instances(row):
    country, code, town = map(lambda idx: row[idx].value, [1, 2, 3])
    if not country:
        raise DataMissingError(fields=['country'])
    if not code:
        raise DataMissingError(fields=['code'])
    if not town:
        raise DataMissingError(fields=['town'])
    return CountryTown(
        country=country,
        town=town,
        code=code
    )


def load_data(file_path, hook, skip_first=True, chunk=2000, sheet=0):
    wb = load_workbook(file_path, read_only=True)
    gen = wb.worksheets[sheet].rows
    if skip_first:
        next(gen)
    for i, rows in enumerate(chunked(gen, chunk)):
        data = list()
        for j, row in enumerate(rows):
            try:
                parsed = hook(row)
                if isinstance(parsed, list):
                    data += parsed
                else:
                    data.append(parsed)
            except (DataMissingError, DataValidationError) as e:
                row_no = i * chunk + j + 1 + int(skip_first)
                print(f'{e} | file={file_path} | row no={row_no}')
        with session_scope() as s:
            s.bulk_save_objects(data)


def encrypt_uid_and_save_file(file_path, column_index, output_dir, skip_first=True, sheet=0):
    wb = load_workbook(file_path, read_only=False)
    gen = wb.worksheets[sheet].rows
    base_name = os.path.basename(file_path)
    output_path = os.path.join(output_dir, base_name)
    if skip_first:
        next(gen)
    for row in gen:
        value = row[column_index].value
        if type(value) == str:
            value = value.strip()
        if validate_uid(value):
            new_uid = 'I-' + blake2b(value.encode('utf-8'), digest_size=5).hexdigest()
        else:
            new_uid = 'INVALID'
        row[column_index].value = new_uid
    wb.save(output_path)
